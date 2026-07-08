from pathlib import Path
import io
import csv
from PyPDF2 import PdfReader

# Maximum product file size: 10 MB
MAX_PRODUCT_FILE_SIZE = 10 * 1024 * 1024  # 10 MB in bytes

# Allowed product file extensions
ALLOWED_PRODUCT_EXTENSIONS = {".xlsx", ".csv", ".xls"}


def get_mime_type(filename: str) -> str:
    ext = Path(filename).suffix.lower()
    mime_map = {
        ".pdf": "application/pdf", ".txt": "text/plain", ".csv": "text/csv",
        ".tsv": "text/tab-separated-values", ".md": "text/markdown",
        ".json": "application/json", ".yaml": "application/x-yaml",
        ".yml": "application/x-yaml", ".xml": "application/xml",
        ".html": "text/html", ".htm": "text/html", ".css": "text/css",
        ".js": "text/javascript", ".ts": "application/typescript",
        ".log": "text/x-log", ".doc": "application/msword",
        ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
    }
    return mime_map.get(ext, "text/plain")


def extract_text_from_bytes(content: bytes, filename: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext == ".txt":
        return content.decode("utf-8", errors="ignore")
    elif ext == ".pdf":
        try:
            reader = PdfReader(io.BytesIO(content))
            return "\n".join([page.extract_text() or "" for page in reader.pages])
        except:
            return ""
    elif ext == ".csv":
        try:
            text_content = content.decode("utf-8", errors="ignore")
            rows = list(csv.reader(io.StringIO(text_content)))
            return "\n".join([", ".join(row) for row in rows])
        except:
            return ""
    return ""


def validate_product_file(filename: str, content_length: int) -> str | None:
    """
    Validate a product file by extension and size.
    Returns an error message string if invalid, or None if valid.
    """
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_PRODUCT_EXTENSIONS:
        return (
            f"Unsupported product file type '{ext}'. "
            f"Allowed: {', '.join(sorted(ALLOWED_PRODUCT_EXTENSIONS))}"
        )
    if content_length > MAX_PRODUCT_FILE_SIZE:
        size_mb = content_length / (1024 * 1024)
        return (
            f"Product file too large ({size_mb:.1f} MB). "
            f"Maximum allowed: {MAX_PRODUCT_FILE_SIZE // (1024 * 1024)} MB"
        )
    return None


def extract_product_file_text(content: bytes, filename: str) -> str:
    """
    Extract structured text from a product file (XLSX, XLS, or CSV).
    Returns a human-readable table representation suitable for prompt injection.

    For XLSX/XLS: processes all sheets, formats each as a named table.
    For CSV: formats as a single table.
    """
    ext = Path(filename).suffix.lower()

    if ext == ".csv":
        return _extract_csv_as_table(content)
    elif ext in (".xlsx", ".xls"):
        return _extract_excel_as_table(content)
    return ""


def _extract_csv_as_table(content: bytes) -> str:
    """Convert CSV bytes to a structured text table."""
    try:
        text_content = content.decode("utf-8", errors="ignore")
        rows = list(csv.reader(io.StringIO(text_content)))
        if not rows:
            return ""

        lines = ["=== Product Data ==="]
        # First row as header
        if len(rows) > 0:
            lines.append(" | ".join(rows[0]))
            lines.append("-" * 60)
        # Data rows
        for row in rows[1:]:
            lines.append(" | ".join(row))

        lines.append(f"\nTotal rows: {len(rows) - 1}")
        return "\n".join(lines)
    except Exception:
        return ""


def _extract_excel_as_table(content: bytes) -> str:
    """
    Convert XLSX/XLS bytes to structured text tables (one per sheet).
    Uses openpyxl for .xlsx files.
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        all_sections = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Convert each cell to string, handle None
                str_row = [str(cell) if cell is not None else "" for cell in row]
                # Skip completely empty rows
                if any(cell.strip() for cell in str_row):
                    rows.append(str_row)

            if not rows:
                continue

            lines = [f"=== Sheet: {sheet_name} ==="]
            # First row as header
            lines.append(" | ".join(rows[0]))
            lines.append("-" * 60)
            # Data rows
            for row in rows[1:]:
                lines.append(" | ".join(row))

            lines.append(f"\nTotal rows in '{sheet_name}': {len(rows) - 1}")
            all_sections.append("\n".join(lines))

        wb.close()
        return "\n\n".join(all_sections)
    except Exception:
        return ""